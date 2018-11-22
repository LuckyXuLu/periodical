# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
import xlwt
# import csv
import re
import json
from openpyxl import load_workbook, Workbook
import pymysql


class SpringerProjectPipeline(object):
    def __init__(self):
        # 打开数据库连接
        self.db = pymysql.connect(host='127.0.0.1', port=3306, user='root',
                                  password='mysql', db='periodical_database', charset='utf8')

        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.db.cursor()

        self.xlsx()
        # self.wb = load_workbook("Springer.xlsx")  # 生成一个已存在的wookbook对象
        # self.ws = self.wb.active  # 激活sheet
        # self.row_n = 2

        self.wb_2 = load_workbook("异常邮箱.xlsx")  # 生成一个已存在的wookbook对象
        self.ws_2 = self.wb_2.active  # 激活sheet
        self.row_n_2 = 2

        self.wb_3 = load_workbook("格式ok异常邮箱.xlsx")  # 生成一个已存在的wookbook对象
        self.ws_3 = self.wb_3.active  # 激活sheet
        self.row_n_3 = 2
        self.cdov = 1
        self.cdov_2 = 1

    def xlsx(self):
        workbook = Workbook()
        booksheet = workbook.active
        booksheet.append(
            ["网站", "出版日期", "学科", "期刊", "姓名", "邮箱", "作者简介", "标题", "关键词", "摘要", "url"])
        # workbook.save('Springer.xlsx')
        workbook.save('异常邮箱.xlsx')
        workbook.save('格式ok异常邮箱.xlsx')

    def open_spider(self, spider):  # 在爬虫启动的时候,仅执行一次
        # 请空数据表
        sql = """truncate table springer_table"""
        # 执行sql语句
        self.cursor.execute(sql)
        # 提交到数据库执行
        self.db.commit()
        pass

    def close_spider(self, spider):  # 在爬虫关闭的时候,仅执行一次
        # sql = """select * from springer_table"""
        # self.cursor.execute(sql.encode('utf-8'))
        # data_list = self.cursor.fetchall()
        # for data in data_list:
        #     row_2 = (data[11], str(data[1]), data[2], data[3], data[4], data[5], data[6], data[9], data[7], data[8], data[10])
        #     try:
        #         for col_n, col in enumerate(row_2):
        #             self.ws.cell(self.row_n, col_n + 1, col)
        #         self.wb.save("Springer.xlsx")  # 保存
        #         self.row_n += 1
        #
        #         print("成功读取: %s个" % self.cdov_2)
        #         self.cdov_2 += 1
        #
        #     except Exception as a:
        #         # for col_n_3, col in enumerate(row_2):
        #         #     self.ws_3.cell(self.row_n_3, col_n_3 + 1, col)
        #         # self.wb_3.save("格式ok异常邮箱.xlsx")  # 保存
        #         # self.row_n_3 += 1
        #         with open("{}.txt".format("格式ok异常邮箱"), "a") as e3:
        #             e3.write("{}*{}*{}*{}*{}*{}*{}*{}*{}*{}*{}\n".format(data[11], str(data[1]), data[2], data[3],
        #             data[4], data[5], data[6], data[9], data[7], data[8].replace("*", "").replace("\\n", "").replace("\n", ""), data[10]))
        #     print("=" * 10)

        # 关闭数据库连接
        self.db.close()
        pass

    def process_item(self, item, spider):
        if (item["resume"] and "\\(\\" in item["resume"]) or (item["标题"] and "\\(\\" in item["标题"]) or \
                (item["关键词"] and "\\(\\" in item["关键词"]):
            self.data_cleaning(item)

        self.data_storage(item)
        # print(json.dumps(item, sort_keys=True, ensure_ascii=False, indent=4))
        return item

    def data_storage(self, item_data):
        if item_data["E_mail"] and item_data["name"]:
            self.data_mali(item_data)
            try:
                row = (
                    item_data["网站"], item_data["出版日期"], item_data["学科"], item_data["期刊"], item_data["name"],
                    item_data["E_mail"], item_data["biography"], item_data["标题"], item_data["关键词"], item_data["resume"], item_data["url"])
                if item_data["判断"]:
                    # for col_n, col in enumerate(row):
                    #     self.ws.cell(self.row_n, col_n + 1, col)
                    # self.wb.save("Springer.xlsx")  # 保存
                    # self.row_n += 1

                    self.add(item_data)  # 存储mysql
                    # print(json.dumps(item_data, sort_keys=True, ensure_ascii=False, indent=4))
                    print("成功ok: %s个" % self.cdov)
                    self.cdov += 1
                else:
                    print("异常邮箱成功ok: %s个" % self.cdov)
                    for col_n_2, col in enumerate(row):
                        self.ws_2.cell(self.row_n_2, col_n_2 + 1, col)
                    self.wb_2.save("异常邮箱.xlsx")  # 保存
                    self.row_n_2 += 1

            except Exception as a:
                print("异常: ", a)
                with open("{}.txt".format("异常"), "a") as e3:
                    e3.write("{}: {}; url: {}\n".format("异常", a, item_data["url"]))
        else:
            with open("{}.txt".format("空邮箱数据"), "a") as e3:
                e3.write("{}url: {}\n".format("空邮箱数据", item_data["url"]))

    # 存储pymysql
    def add(self, item_data):
        if not item_data["出版日期"]:
            item_data["出版日期"] = 2010
        sql = """insert into springer_table values(id,
                          {},
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          "{}",
                          default)""".format(int(item_data["出版日期"]), item_data["学科"], item_data["期刊"],
                                             item_data["name"], item_data["E_mail"], item_data["biography"],
                                             item_data["关键词"], item_data["resume"], item_data["标题"],
                                             item_data["url"], item_data["网站"])
        try:
            # 执行sql语句
            self.cursor.execute(sql)
            # 提交到数据库执行
            self.db.commit()
        except Exception as a:
            row_3 = (item_data["网站"], item_data["出版日期"], item_data["学科"], item_data["期刊"], item_data["name"],
                    item_data["E_mail"], item_data["biography"], item_data["标题"], item_data["关键词"], item_data["resume"], item_data["url"])
            print("row_3: ", row_3)
            for col_n_3, col in enumerate(row_3):
                self.ws_3.cell(self.row_n_3, col_n_3 + 1, col)
            self.wb_3.save("格式ok异常邮箱.xlsx")  # 保存
            self.row_n_3 += 1

            # with open("{}.txt".format("格式ok异常邮箱"), "a") as e3:
            #     e3.write("{}*{}*{}*{}*{}*{}*{}*{}*{}*{}*{}\n".format(item_data["网站"], item_data["出版日期"], item_data["学科"], item_data["期刊"], item_data["name"], item_data["E_mail"], item_data["biography"], item_data["标题"], item_data["关键词"], item_data["resume"], item_data["url"]))
            # 如果发生错误则回滚
            self.db.rollback()

    def data_cleaning(self, item_data):
        re_gular = re.compile(r'\\\(.+?\\\)')
        if item_data["标题"]:
            headline_list = re_gular.findall(item_data["标题"])
            for headline in headline_list:
                headline_str_1 = str()
                for i in headline.split(">"):
                    re_gular_2 = re.compile(r'\{(.*?)\}')
                    headline_list_2 = re_gular_2.findall(i)

                    headline_str_2 = str()
                    for headline_2 in headline_list_2:
                        headline_str_2 += headline_2

                    if headline_str_1:
                        headline_str_1 += ">" + " ".join(headline_str_2.split())
                    else:
                        headline_str_1 += " ".join(headline_str_2.split())
                headline_str_3 = headline_str_1.replace("{", "")
                item_data["标题"] = item_data["标题"].replace(headline, headline_str_3)

        if item_data["resume"]:
            digest_list = re_gular.findall(item_data["resume"])
            for digest in digest_list:
                digest_str_1 = str()
                for i in digest.split(">"):
                    re_gular_2 = re.compile(r'\{(.*?)\}')
                    digest_list_2 = re_gular_2.findall(i)

                    digest_str_2 = str()
                    for digest_2 in digest_list_2:
                        digest_str_2 += digest_2

                    if digest_str_1:
                        digest_str_1 += ">" + " ".join(digest_str_2.split())
                    else:
                        digest_str_1 += " ".join(digest_str_2.split())
                digest_str_3 = digest_str_1.replace("{", "")
                item_data["标题"] = item_data["标题"].replace(digest, digest_str_3)

        if item_data["关键词"]:
            headline_list = re_gular.findall(item_data["关键词"])
            for headline in headline_list:
                headline_str_1 = str()
                for i in headline.split(">"):
                    re_gular_2 = re.compile(r'\{(.*?)\}')
                    headline_list_2 = re_gular_2.findall(i)

                    headline_str_2 = str()
                    for headline_2 in headline_list_2:
                        headline_str_2 += headline_2

                    if headline_str_1:
                        headline_str_1 += ">" + " ".join(headline_str_2.split())
                    else:
                        headline_str_1 += " ".join(headline_str_2.split())
                headline_str_3 = headline_str_1.replace("{", "")
                item_data["关键词"] = item_data["关键词"].replace(headline, headline_str_3)

    def data_mali(self, item_data):
        email = item_data["E_mail"].replace("(", "").replace(")", "").replace("email:", "")
        email = " ".join(email.split())
        if re.match('^[A-Za-z0-9](([A-Za-z0-9\-\_]{1,}(\.[A-Za-z0-9][A-Za-z0-9\-\_]{0,}){0,})|([A-Za-z0-9\-\_]{0,}(\.[A-Za-z0-9][A-Za-z0-9\-\_]{0,}){1,}))@(((([A-Za-z][A-Za-z\-\_]{1,}(\.[A-Za-z][A-Za-z\-\_]{1,}){0,})|([A-Za-z][A-Za-z\-\_]{0,}(\.[A-Za-z][A-Za-z\-\_]{1,}){1,})|(163)|(126)|(263)|(139)|(xs4all)|(iut-tlse3)|(unina2)|(uniroma1)|(uniroma3)|(jasatirta1))\.(([A-z]{2,3})|(info)|(coop))$)|([A-z0-9][A-z0-9\.\-]{0,}[A-z0-9]{1,}\.jp$)|(vip\.163\.com$))', email):  # 判断字符串是否为字母
            pass
        else:
            item_data["判断"] = None

