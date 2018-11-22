# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import load_workbook, Workbook
import json
import re
import pymysql


class LetpubProjectPipeline(object):
    def __init__(self):
        # # 打开数据库连接
        # self.db = pymysql.connect(host='127.0.0.1', port=3306, user='root',
        #                           password='mysql', db='tandfonline_database', charset='utf8')
        #
        # # 使用 cursor() 方法创建一个游标对象 cursor
        # self.cursor = self.db.cursor()

        self.xlsx()
        self.wb = load_workbook("letpub.xlsx")  # 生成一个已存在的wookbook对象
        self.ws = self.wb.active  # 激活sheet
        self.row_n = 2

        self.cdov = 1

    def open_spider(self, spider):  # 在爬虫启动的时候,仅执行一次
        pass

    def close_spider(self, spider):  # 在爬虫关闭的时候,仅执行一次
        # 关闭数据库连接
        # self.db.close()
        pass

    def xlsx(self):
        workbook = Workbook()
        booksheet = workbook.active
        booksheet.append(
            ["网站", "学科", "ISSN", "期刊", "影响因子", "中科院分区", "大类学科", "小类学科", "SCI/SCIE", "是否OA", "录用比例",
             "审稿周期", "查看数", "自引率", "五年影响因子", "期刊官方网站", "期刊投稿网址", "通讯方式", "涉及的研究方向",
             "出版国家或地区", "出版周期", "出版年份", "年文章数", "url"])
        workbook.save('letpub.xlsx')

    def process_item(self, item, spider):

        row = (
            item["网站"], item["学科"], item["ISSN"], item["期刊"], item["影响因子"], item["中科院分区"], item["大类学科"],
            item["小类学科"], item["SCI/SCIE"], item["是否OA"], item["录用比例"], item["审稿周期"], item["查看数"],
            item["自引率"], item["五年影响因子"], item["期刊官方网站"], item["期刊投稿网址"], item["通讯方式"], item["涉及的研究方向"],
            item["出版国家或地区"], item["出版周期"], item["出版年份"], item["年文章数"], item["url"])

        try:
            for col_n, col in enumerate(row):
                self.ws.cell(self.row_n, col_n + 1, col)
            self.wb.save("letpub.xlsx")  # 保存
            self.row_n += 1

            # self.add(item)  # 存储mysql
            print(json.dumps(item, sort_keys=True, ensure_ascii=False, indent=4))
            print("成功ok: %s个" % self.cdov)
            self.cdov += 1
        except Exception as a:
            print("异常: ", a)

        return item
