import pymysql
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl import load_workbook, Workbook
import time


class Mysql_Ob(object):
    def __init__(self,  database, data_table):
        # 打开数据库连接
        self.db = pymysql.connect(host='127.0.0.1', port=3306, user='root',
                                  password='mysql', db=database, charset='utf8')

        # 使用 cursor() 方法创建一个游标对象 cursor
        self.cursor = self.db.cursor()
        self.cdov = 1

        # 数据表
        self.data_table = data_table

        self.xlsx()
        self.wb = load_workbook("Springer1.xlsx")  # 生成一个已存在的wookbook对象
        self.ws = self.wb.active  # 激活sheet
        self.row_n = 2

    def xlsx(self):
        workbook = Workbook()
        booksheet = workbook.active
        booksheet.append(
            ["网站", "出版日期", "学科", "期刊", "姓名", "邮箱", "作者简介", "标题", "关键词", "摘要", "url"])
        workbook.save('Springer1.xlsx')


    # 创建表学科
    def create_table(self):
        # 使用预处理语句创建表

        sql = """create table {} (id int unsigned not null auto_increment primary key,
                出版日期 int(4),
                学科 varchar(500) not null,
                期刊 varchar(500) not null,
                name varchar(100) not null,
                email varchar(100) not null,
                作者简介 varchar(100) null,
                关键词 varchar(1000) null,
                摘要 varchar(10000) null,
                标题 varchar(1000) null,
                url varchar(500) not null,
                网站 varchar(500) not null,
                更新时间 TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP)""".format(self.data_table)
        self.cursor.execute(sql)

    # 查询读取数据
    def inquire_data(self):
        oi = 1
        id_codv = 1
        id_codv_2 = id_codv + 2000
        while True:
            sql = """select * from {} where id={}""".format(self.data_table, id_codv)
            self.cursor.execute(sql.encode('utf-8'))
            data_list = self.cursor.fetchall()
            if not data_list:
                print("等待10秒...")
                time.sleep(10)
                continue
            elif id_codv == id_codv_2:
                # return id_codv
                id_codv_2 += 2000
                oi += 1

                workbook = Workbook()
                booksheet = workbook.active
                booksheet.append(
                    ["网站", "出版日期", "学科", "期刊", "姓名", "邮箱", "作者简介", "标题", "关键词", "摘要", "url"])
                workbook.save('Springer{}.xlsx'.format(oi))

                self.wb = load_workbook("Springer{}.xlsx".format(oi))  # 生成一个已存在的wookbook对象
                self.ws = self.wb.active  # 激活sheet
                self.row_n = 2

            id_codv += 1
            for data in data_list:
                row_2 = (data[11], str(data[1]), data[2], data[3], data[4], data[5], data[6], data[9], data[7], data[8], data[10])
                try:
                    for col_n, col in enumerate(row_2):
                        self.ws.cell(self.row_n, col_n + 1, col)
                    self.wb.save("Springer{}.xlsx".format(oi))  # 保存
                    self.row_n += 1

                    print("第{}个, 成功读取".format(self.cdov))
                    self.cdov += 1
                except Exception as a:
                    with open("{}.txt".format("格式ok异常邮箱{}".format(oi)), "a") as e3:
                        e3.write("{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}\n".format(data[11], str(data[1]), data[2], data[3], data[4],
                                                                             data[5], data[6], data[9], data[7], data[8], data[10]))

    # 写入数据
    def write_data(self):
        workbook = load_workbook('Springer.xlsx')
        booksheet = workbook.active
        rows = booksheet.rows
        # 迭代所有的行
        for row in rows:
            line = [col.value for col in row]
            if not line[1]:
                line[1] = 1001

            sql = """insert into subject values(id,
                              {},
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              "{}",
                              default)""".format(int(line[1]), line[2], line[3], line[5],
                                                 line[6], line[7], line[8], line[9], line[10], line[11], line[12],
                                                 line[13], line[14], line[15], line[16], line[17], line[18], line[4],
                                                 line[19], line[0])
            try:
                # 执行sql语句
                self.cursor.execute(sql)
                # 提交到数据库执行
                self.db.commit()
                print("ok")
                print("第{}个, 成功".format(self.cdov))
                self.cdov += 1
            except Exception as a:
                # 如果发生错误则回滚
                print("NG" * 20)
                with open("异常数据.txt", 'a') as e:
                    e.write("第{}个, 异常\n".format(self.cdov))
                self.cdov += 1
                self.db.rollback()
        print("==" * 100)

    def mia(self):
        # 创建数据表结构
        if False:
            self.create_table()
            print("表创建OK")

        # 查询数据
        if True:
            id_str = self.inquire_data()
            print("查询完成 {}".format(id_str))

        # 写入
        if False:
            self.write_data()
            # 关闭数据库连接
            self.db.close()
            print("数据写入OK")


if __name__ == '__main__':
    database = 'periodical_database'
    data_table = 'springer_table'
    mysql = Mysql_Ob(database, data_table)
    mysql.mia()
