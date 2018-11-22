# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://doc.scrapy.org/en/latest/topics/item-pipeline.html
from openpyxl import load_workbook, Workbook
import json
import re
import pymysql


class TandfonlineProjectPipeline(object):
    def __init__(self):
        # # 打开数据库连接
        # self.db = pymysql.connect(host='127.0.0.1', port=3306, user='root',
        #                           password='mysql', db='tandfonline_database', charset='utf8')
        #
        # # 使用 cursor() 方法创建一个游标对象 cursor
        # self.cursor = self.db.cursor()

        self.xlsx()
        self.wb = load_workbook("tandfonline.xlsx")  # 生成一个已存在的wookbook对象
        self.ws = self.wb.active  # 激活sheet
        self.row_n = 2

        self.wb_2 = load_workbook("异常邮箱.xlsx")  # 生成一个已存在的wookbook对象
        self.ws_2 = self.wb_2.active  # 激活sheet
        self.row_n_2 = 2

        self.cdov = 1

    def open_spider(self, spider):  # 在爬虫启动的时候,仅执行一次
        pass

    def close_spider(self, spider):  # 在爬虫关闭的时候,仅执行一次
        # 关闭数据库连接
        # self.db.close()
        pass

    def xlsx(self):
        workbook = Workbook()
        booksheet = workbook.active
        booksheet.append(
            ["网站", "出版日期", "学科", "期刊", "标题", "姓名", "邮箱", "作者简介", "关键词", "摘要", "url"])
        workbook.save('tandfonline.xlsx')
        workbook.save('异常邮箱.xlsx')

    def process_item(self, item, spider):

        if item["邮箱"]:
            self.data_mali(item)
            row = (
                item["网站"], item["出版日期"], item["学科"], item["期刊"], item["标题"],
                item["姓名"], item["邮箱"], item["作者简介"], item["关键词"], item["摘要"], item["url"])

            try:
                if item["判断"]:
                    for col_n, col in enumerate(row):
                        self.ws.cell(self.row_n, col_n + 1, col)
                    self.wb.save("tandfonline.xlsx")  # 保存
                    self.row_n += 1

                    # self.add(item)  # 存储mysql
                    # print(json.dumps(item_data, sort_keys=True, ensure_ascii=False, indent=4))
                    print("成功ok: %s个" % self.cdov)
                    self.cdov += 1
                else:
                    for col_n_2, col in enumerate(row):
                        self.ws_2.cell(self.row_n_2, col_n_2 + 1, col)
                    self.wb_2.save("异常邮箱.xlsx")  # 保存
                    self.row_n_2 += 1
            except Exception as a:
                print("异常: ", a)
        else:
            with open("异常数据.txt", "a") as e3:
                e3.write("异常数据: {}\n".format(item))

        return item

    # def add(self, item_data):
    #     if not item_data["出版日期"]:
    #         item_data["出版日期"] = 1001
    #     sql = """insert into subject values(id,
    #                       {},
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       "{}",
    #                       default)""".format(int(item_data["出版日期"]), item_data["学科"], item_data["期刊"], item_data["姓名"],
    #                                          item_data["邮箱"], item_data["作者简介"], item_data["关键词"],
    #                                          item_data["摘要"], item_data["标题"], item_data["url"], item_data["网站"])
    #     try:
    #         # 执行sql语句
    #         self.cursor.execute(sql)
    #         # 提交到数据库执行
    #         self.db.commit()
    #     except Exception as a:
    #         # 如果发生错误则回滚
    #         self.db.rollback()

    def data_mali(self, item_data):

            mail = item_data["邮箱"].replace("(", "").replace(")", "").replace("email:", "")
            mail = " ".join(mail.split())

            if re.match('^[A-Za-z0-9](([A-Za-z0-9\-\_]{1,}(\.[A-Za-z0-9][A-Za-z0-9\-\_]{0,}){0,})|([A-Za-z0-9\-\_]{0,}(\.[A-Za-z0-9][A-Za-z0-9\-\_]{0,}){1,}))@(((([A-Za-z][A-Za-z\-\_]{1,}(\.[A-Za-z][A-Za-z\-\_]{1,}){0,})|([A-Za-z][A-Za-z\-\_]{0,}(\.[A-Za-z][A-Za-z\-\_]{1,}){1,})|(163)|(126)|(263)|(139)|(xs4all)|(iut-tlse3)|(unina2)|(uniroma1)|(uniroma3)|(jasatirta1))\.(([A-z]{2,3})|(info)|(coop))$)|([A-z0-9][A-z0-9\.\-]{0,}[A-z0-9]{1,}\.jp$)|(vip\.163\.com$))', mail):  # 判断字符串是否为字母
                pass
            else:
                item_data["判断"] = None
